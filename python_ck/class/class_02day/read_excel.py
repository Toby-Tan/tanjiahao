from openpyxl import load_workbook


class Rexcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []  # 数据存储为列表格式
        print(sheet.max_row)
        for x in range(1, sheet.max_row):
            sub_data = {'url': sheet.cell(x + 1, 2).value, 'data': sheet.cell(x + 1, 3).value,
                        'method': sheet.cell(x + 1, 4).value}  # 数据存储为字典格式

            test_data.append(sub_data)
        return test_data  # 返回获取到的数据


if __name__ == '__main__':
    re = Rexcel('data.xlsx', 'Sheet1')
    data = re.get_data()
    for row in data:
        print(row)

